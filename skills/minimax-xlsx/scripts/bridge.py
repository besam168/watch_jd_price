import argparse
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook


def read_json(path: str):
    return json.loads(Path(path).read_text(encoding='utf-8'))


def ensure_parent(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)


def create_workbook(spec: dict, out_path: Path):
    wb = Workbook()
    default_ws = wb.active
    first = True

    for sheet_spec in spec.get('sheets', []):
        title = sheet_spec.get('name', 'Sheet1')
        if first:
            ws = default_ws
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)

        rows = sheet_spec.get('rows', [])
        for row in rows:
            ws.append(row)

        widths = sheet_spec.get('widths', {})
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    if first:
        default_ws.title = 'Sheet1'

    ensure_parent(out_path)
    wb.save(out_path)


def analyze_workbook(in_path: Path):
    wb = load_workbook(in_path, data_only=False)
    result = {
        'path': str(in_path),
        'sheet_names': wb.sheetnames,
        'sheets': []
    }
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        preview = rows[:5]
        result['sheets'].append({
            'name': ws.title,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'preview': preview,
        })
    return result


def update_cells(spec: dict, in_path: Path, out_path: Path | None):
    wb = load_workbook(in_path)
    for op in spec.get('updates', []):
        sheet = op['sheet']
        cell = op['cell']
        value = op.get('value')
        ws = wb[sheet]
        ws[cell] = value
    target = out_path or in_path
    ensure_parent(target)
    wb.save(target)
    return str(target)


def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest='cmd', required=True)

    p_create = sub.add_parser('create')
    p_create.add_argument('--input', required=True)
    p_create.add_argument('--output', required=True)

    p_analyze = sub.add_parser('analyze')
    p_analyze.add_argument('--input', required=True)

    p_update = sub.add_parser('update-cells')
    p_update.add_argument('--input', required=True, help='workbook path')
    p_update.add_argument('--spec', required=True, help='json update spec path')
    p_update.add_argument('--output', required=False)

    args = parser.parse_args()

    if args.cmd == 'create':
        spec = read_json(args.input)
        create_workbook(spec, Path(args.output))
        print(args.output)
    elif args.cmd == 'analyze':
        result = analyze_workbook(Path(args.input))
        print(json.dumps(result, ensure_ascii=False, indent=2))
    elif args.cmd == 'update-cells':
        spec = read_json(args.spec)
        out = update_cells(spec, Path(args.input), Path(args.output) if args.output else None)
        print(out)


if __name__ == '__main__':
    main()
