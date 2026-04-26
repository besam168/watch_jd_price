from openpyxl import Workbook

out = r"C:\Users\besam\.openclaw\workspace\_tmp\xlsx-smoke-test.xlsx"
wb = Workbook()
ws = wb.active
ws.title = 'Sheet1'
ws['A1'] = 'Item'
ws['B1'] = 'Amount'
ws['A2'] = 'Alpha'
ws['B2'] = 12
ws['A3'] = 'Beta'
ws['B3'] = 30
ws['A4'] = 'Total'
ws['B4'] = '=SUM(B2:B3)'
wb.save(out)
print(out)
